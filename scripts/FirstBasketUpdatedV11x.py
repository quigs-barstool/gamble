from nba_api.stats.endpoints import LeagueGameFinder, ScoreboardV2
from nba_api.stats.library.parameters import SeasonType

import pandas as pd
import numpy as np
import datetime
import time
import re
import os
import random
import requests
from requests.exceptions import ReadTimeout, ConnectionError, Timeout

SEASON = "2025-26"
SEASON_TYPE = SeasonType.regular

# Throttles
SLEEP_BETWEEN_CALLS = 0.55          # between PBP calls when backfilling new games
SCOREBOARD_SLEEP = 0.25            # between ScoreboardV2 calls for ThisWeeksMatchups

# Cache file: per-team per-game results so we only fetch NEW games next time
CACHE_FILE = f"first_shot_cache_{SEASON.replace('-', '')}.csv"

# Sessions (persistent)
SESSION = requests.Session()
CDN_SESSION = requests.Session()

# -----------------------------
# Headers for stats.nba.com fallback
# -----------------------------
NBA_HEADERS = {
    "Host": "stats.nba.com",
    "Connection": "keep-alive",
    "Accept": "application/json, text/plain, */*",
    "x-nba-stats-token": "true",
    "x-nba-stats-origin": "stats",
    "Origin": "https://www.nba.com",
    "Referer": "https://www.nba.com/",
    "Accept-Language": "en-US,en;q=0.9",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
}


# ---------------------------------------------------------
#   GENERIC RETRY WRAPPER
# ---------------------------------------------------------
def with_retries(fn, label: str, max_retries: int = 5, base_sleep: float = 2.0):
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            return fn()
        except (ReadTimeout, ConnectionError, TimeoutError, Timeout) as e:
            last_err = e
            sleep_s = base_sleep * (2 ** (attempt - 1)) + random.uniform(0.0, 0.6)
            print(f"  {label} timeout (attempt {attempt}/{max_retries}) — sleeping {sleep_s:.1f}s then retrying...")
            time.sleep(sleep_s)
    raise last_err


# ---------------------------------------------------------
#   HELPERS
# ---------------------------------------------------------
def is_gleague_game_id(game_id: str) -> bool:
    return str(game_id).startswith("20")


def round2(df: pd.DataFrame, cols):
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
    return df


def _looks_like_html(text: str) -> bool:
    if not isinstance(text, str):
        return False
    t = text.strip().lower()
    return t.startswith("<!doctype html") or t.startswith("<html") or "<head" in t[:500] or "<body" in t[:500]


def sec_remaining_from_clock(clock_val) -> int | None:
    """
    CDN liveData clock commonly looks like:
      - "PT11M34.00S" (ISO-ish duration remaining)
    Sometimes you may see "11:34".
    Return seconds remaining in the period.
    """
    if clock_val is None:
        return None
    s = str(clock_val).strip()

    # Case: "11:34"
    if re.match(r"^\d{1,2}:\d{2}$", s):
        m, ss = s.split(":")
        try:
            return int(m) * 60 + int(ss)
        except Exception:
            return None

    # Case: "PT11M34.00S"
    if s.startswith("PT"):
        mm = re.search(r"PT(\d+)M", s)
        ss = re.search(r"M(\d+)(?:\.\d+)?S", s)
        try:
            m = int(mm.group(1)) if mm else 0
            sec = int(ss.group(1)) if ss else 0
            return m * 60 + sec
        except Exception:
            return None

    return None


def safe_upper(x) -> str:
    return str(x).upper() if isinstance(x, str) else ""


def dedupe_game_ids(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "GAME_ID" not in df.columns:
        return df
    out = df.copy()
    out["GAME_ID"] = out["GAME_ID"].astype(str)
    out = out.drop_duplicates(subset=["GAME_ID"], keep="first").copy()
    return out


# ---------------------------------------------------------
#   BINNED WIN RATE + WIN PROBABILITY MAPPING
# ---------------------------------------------------------
def compute_binned_win_rates(
    hist_df: pd.DataFrame,
    score_col: str = "Combined_QUIGS_POWER_SCORE",
    outcome_col: str = "Win/Loss",
    bins: int = 10,
) -> tuple[pd.DataFrame, np.ndarray, list[str]]:
    """
    Build a binned win rate table from historical games.

    - Uses quantile bins (pd.qcut) so bins have ~equal counts.
    - Ignores outcome 'NA' rows.
    - Returns: (binned_table_df, bin_edges, labels)
    """
    if hist_df is None or hist_df.empty:
        empty = pd.DataFrame(columns=["Bin", "Bin_Low", "Bin_High", "Bin_Mid", "Games", "Wins", "Win_Rate"])
        return empty, np.array([]), []

    df = hist_df.copy()
    df[score_col] = pd.to_numeric(df[score_col], errors="coerce")
    df = df[df[score_col].notna()].copy()
    df = df[df[outcome_col].isin(["Win", "Loss"])].copy()

    if df.empty or df[score_col].nunique() < 2:
        empty = pd.DataFrame(columns=["Bin", "Bin_Low", "Bin_High", "Bin_Mid", "Games", "Wins", "Win_Rate"])
        return empty, np.array([]), []

    # qcut -> edges
    try:
        cats, edges = pd.qcut(df[score_col], q=bins, retbins=True, duplicates="drop")
    except Exception:
        empty = pd.DataFrame(columns=["Bin", "Bin_Low", "Bin_High", "Bin_Mid", "Games", "Wins", "Win_Rate"])
        return empty, np.array([]), []

    # Create readable labels from edges
    edges = np.array(edges, dtype=float)
    labels = []
    for i in range(len(edges) - 1):
        labels.append(f"{edges[i]:.2f} to {edges[i+1]:.2f}")

    df["BIN_LABEL"] = pd.cut(df[score_col], bins=edges, labels=_toggle_labels(labels), include_lowest=True)

    # If pd.cut produced all NaN due to edge issues, fallback to Interval string
    if df["BIN_LABEL"].isna().all():
        df["BIN_LABEL"] = cats.astype(str)

    df["IS_WIN"] = (df[outcome_col] == "Win").astype(int)

    grp = df.groupby("BIN_LABEL", dropna=True).agg(
        Games=(score_col, "size"),
        Wins=("IS_WIN", "sum"),
        Avg_Score=(score_col, "mean"),
    ).reset_index()

    grp["Win_Rate"] = grp["Wins"] / grp["Games"].replace({0: np.nan})
    grp["Win_Rate"] = grp["Win_Rate"].fillna(0)

    # Bin low/high/mid from label when possible
    bin_low = []
    bin_high = []
    for v in grp["BIN_LABEL"].astype(str).tolist():
        m = re.findall(r"-?\d+\.\d+|-?\d+", v)
        if len(m) >= 2:
            lo = float(m[0])
            hi = float(m[1])
        else:
            lo, hi = np.nan, np.nan
        bin_low.append(lo)
        bin_high.append(hi)

    grp.rename(columns={"BIN_LABEL": "Bin"}, inplace=True)
    grp["Bin_Low"] = bin_low
    grp["Bin_High"] = bin_high
    grp["Bin_Mid"] = (grp["Bin_Low"] + grp["Bin_High"]) / 2
    grp["Bin_Mid"] = grp["Bin_Mid"].fillna(grp["Avg_Score"])

    # Order by mid
    grp = grp.sort_values("Bin_Mid", ascending=True).reset_index(drop=True)

    out = grp[["Bin", "Bin_Low", "Bin_High", "Bin_Mid", "Games", "Wins", "Win_Rate"]].copy()
    return out, edges, labels


def _toggle_labels(labels: list[str]) -> list[str]:
    # Helper: pd.cut requires labels length == len(edges)-1
    return labels


def add_win_probabilities(
    df_matchups: pd.DataFrame,
    binned_table: pd.DataFrame,
    edges: np.ndarray,
    score_col: str = "Combined_QUIGS_POWER_SCORE",
    prob_col: str = "Win_Prob",
) -> pd.DataFrame:
    """
    Add Win_Prob to matchups using the binned win rates.
    """
    if df_matchups is None or df_matchups.empty:
        return df_matchups

    out = df_matchups.copy()
    out[score_col] = pd.to_numeric(out[score_col], errors="coerce")

    if binned_table is None or binned_table.empty or edges is None or len(edges) < 2:
        out[prob_col] = pd.NA
        return out

    # map by bin label
    # Build labels from edges the same way
    labels = [f"{edges[i]:.2f} to {edges[i+1]:.2f}" for i in range(len(edges) - 1)]
    bin_labels = pd.cut(out[score_col], bins=edges, labels=labels, include_lowest=True)

    rate_map = dict(zip(binned_table["Bin"].astype(str), binned_table["Win_Rate"]))
    out[prob_col] = bin_labels.astype(str).map(rate_map)

    # If score is NaN or falls outside bins, leave NA
    out.loc[out[score_col].isna(), prob_col] = pd.NA

    # nice rounding for display
    out[prob_col] = pd.to_numeric(out[prob_col], errors="coerce").round(3)
    return out


# ---------------------------------------------------------
#   GET ALL GAMES FOR SEASON
# ---------------------------------------------------------
def get_games_for_season(season: str, season_type: str = SeasonType.regular) -> pd.DataFrame:
    lgf = with_retries(
        lambda: LeagueGameFinder(season_nullable=season, season_type_nullable=season_type),
        label="LeagueGameFinder",
        max_retries=4,
        base_sleep=2.0
    )
    games = lgf.get_data_frames()[0]

    cols_to_keep = ["GAME_ID", "TEAM_ID", "TEAM_ABBREVIATION", "GAME_DATE", "MATCHUP"]
    games = games[cols_to_keep].drop_duplicates(subset=["GAME_ID", "TEAM_ID"]).copy()
    games["GAME_ID"] = games["GAME_ID"].astype(str)

    # Ignore G League games
    games = games[~games["GAME_ID"].str.startswith("20")].copy()
    return games


# ---------------------------------------------------------
#   PLAYBYPLAY FETCH (CDN FIRST, stats.nba.com FALLBACK)
# ---------------------------------------------------------
def fetch_playbyplay_df(game_id: str) -> pd.DataFrame | None:
    game_id = str(game_id)
    if is_gleague_game_id(game_id):
        return None

    cdn_url = f"https://cdn.nba.com/static/json/liveData/playbyplay/playbyplay_{game_id}.json"

    def cdn_call():
        r = CDN_SESSION.get(cdn_url, timeout=20)
        r.raise_for_status()
        return r

    try:
        r = with_retries(cdn_call, label=f"CDN PBP {game_id}", max_retries=3, base_sleep=1.2)
        raw = r.json()
        actions = (((raw or {}).get("game") or {}).get("actions")) or []
        if not actions:
            print(f"  ❌ Skipping PlayByPlay for game {game_id}: CDN returned no actions")
            return None

        df = pd.DataFrame(actions)

        if "period" not in df.columns or "clock" not in df.columns:
            print(f"  ❌ Skipping PlayByPlay for game {game_id}: CDN format missing period/clock")
            return None

        rename_map = {
            "period": "PERIOD",
            "clock": "CLOCK",
            "teamId": "TEAM_ID",
            "description": "DESCRIPTION",
            "actionType": "ACTION_TYPE",
            "shotValue": "SHOT_VALUE",
            "shotResult": "SHOT_RESULT",
            "subType": "SUB_TYPE",
            "actionNumber": "actionNumber",
        }
        for k, v in rename_map.items():
            if k in df.columns:
                df.rename(columns={k: v}, inplace=True)

        if "TEAM_ID" in df.columns:
            df["TEAM_ID"] = pd.to_numeric(df["TEAM_ID"], errors="coerce")

        df["SEC_REMAINING"] = df["CLOCK"].apply(sec_remaining_from_clock)

        if "DESCRIPTION" in df.columns:
            df["DESC_UPPER"] = df["DESCRIPTION"].apply(safe_upper)
        else:
            df["DESC_UPPER"] = ""

        if "actionNumber" in df.columns:
            df["actionNumber"] = pd.to_numeric(df["actionNumber"], errors="coerce")
            df.sort_values(["PERIOD", "actionNumber"], inplace=True)

        return df

    except Exception as e:
        print(f"  ⚠️ CDN PBP failed for {game_id}: {e}. Trying stats.nba.com fallback...")

    url = "https://stats.nba.com/stats/playbyplayv2"
    params = {"GameID": game_id, "StartPeriod": 0, "EndPeriod": 14}

    def stats_call():
        rr = SESSION.get(url, params=params, headers=NBA_HEADERS, timeout=60)
        rr.raise_for_status()
        return rr

    try:
        rr = with_retries(stats_call, label=f"stats.nba.com PBP {game_id}", max_retries=3, base_sleep=2.0)
        txt = rr.text or ""
        if _looks_like_html(txt):
            print(f"  ❌ Skipping PlayByPlay for game {game_id}: stats.nba.com returned HTML/blocked")
            return None

        raw = rr.json()

        rs = None
        if isinstance(raw, dict):
            if "resultSets" in raw and isinstance(raw["resultSets"], list) and raw["resultSets"]:
                rs = raw["resultSets"][0]
            elif "resultSet" in raw and isinstance(raw["resultSet"], dict):
                rs = raw["resultSet"]

        if not rs or "headers" not in rs or "rowSet" not in rs:
            print(f"  ❌ Skipping PlayByPlay for game {game_id}: unexpected stats.nba.com payload")
            return None

        df = pd.DataFrame(rs["rowSet"], columns=rs["headers"])
        if df.empty:
            print(f"  ❌ Skipping PlayByPlay for game {game_id}: empty stats.nba.com PBP")
            return None

        if "PCTIMESTRING" in df.columns:
            df["SEC_REMAINING"] = df["PCTIMESTRING"].apply(lambda s: sec_remaining_from_clock(s))
        else:
            df["SEC_REMAINING"] = None

        def desc_u(row):
            for c in ("HOMEDESCRIPTION", "VISITORDESCRIPTION", "NEUTRALDESCRIPTION"):
                v = row.get(c, "")
                if isinstance(v, str) and v.strip():
                    return v.upper()
            return ""
        df["DESC_UPPER"] = df.apply(desc_u, axis=1)

        for c in ("EVENTMSGTYPE", "PERIOD", "EVENTNUM", "PLAYER1_TEAM_ID"):
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

        return df

    except (ReadTimeout, ConnectionError, TimeoutError, Timeout) as e:
        print(f"  ❌ Skipping PlayByPlay for game {game_id} (stats fallback timed out): {e}")
        return None
    except Exception as e:
        print(f"  ❌ Skipping PlayByPlay for game {game_id} (stats fallback failed): {e}")
        return None


# ---------------------------------------------------------
#   COMPUTE FIRST MINUTE POINTS
# ---------------------------------------------------------
def compute_first_minute_points(df_pbp: pd.DataFrame, home_team_id: int, away_team_id: int) -> dict:
    home_team_id = int(home_team_id)
    away_team_id = int(away_team_id)
    points_by_team = {home_team_id: 0, away_team_id: 0}

    if df_pbp is None or df_pbp.empty:
        return points_by_team

    if {"TEAM_ID", "PERIOD", "SEC_REMAINING"}.issubset(df_pbp.columns):
        df = df_pbp.copy()

        fm = df[
            (df["PERIOD"] == 1) &
            (df["SEC_REMAINING"].notna()) &
            (df["SEC_REMAINING"] > 11 * 60) &
            (df["SEC_REMAINING"] <= 12 * 60)
        ].copy()

        if fm.empty:
            return points_by_team

        if "SHOT_RESULT" in fm.columns:
            fm["SHOT_RESULT"] = fm["SHOT_RESULT"].astype(str)
        else:
            fm["SHOT_RESULT"] = ""

        def is_made_shot(row) -> bool:
            sr = str(row.get("SHOT_RESULT", "")).lower()
            if sr in ("made", "make", "1"):
                return True
            if sr in ("missed", "miss", "0"):
                return False
            return "MISS" not in str(row.get("DESC_UPPER", ""))

        def is_shot_action(row) -> bool:
            if pd.notna(row.get("SHOT_VALUE", None)):
                return True
            at = str(row.get("ACTION_TYPE", "")).lower()
            return "shot" in at or at in ("2pt", "3pt", "jump_shot", "layup", "dunk")

        shots = fm[fm.apply(is_shot_action, axis=1)].copy()
        if not shots.empty:
            def shot_val(row) -> int:
                sv = row.get("SHOT_VALUE", None)
                if pd.notna(sv):
                    try:
                        return int(sv)
                    except Exception:
                        pass
                d = str(row.get("DESC_UPPER", ""))
                if "3PT" in d or "3-PT" in d:
                    return 3
                return 2

            shots["PTS"] = shots.apply(lambda r: shot_val(r) if is_made_shot(r) else 0, axis=1)
            spts = shots.groupby("TEAM_ID")["PTS"].sum().to_dict()
            for tid, pts in spts.items():
                if int(tid) in points_by_team:
                    points_by_team[int(tid)] += int(pts)

        def is_ft(row) -> bool:
            if "FREE THROW" in str(row.get("DESC_UPPER", "")):
                return True
            at = str(row.get("ACTION_TYPE", "")).lower()
            return "free" in at and "throw" in at

        ft = fm[fm.apply(is_ft, axis=1)].copy()
        if not ft.empty:
            made = ft[~ft["DESC_UPPER"].str.contains("MISS", na=False)]
            counts = made.groupby("TEAM_ID").size().to_dict()
            for tid, cnt in counts.items():
                if int(tid) in points_by_team:
                    points_by_team[int(tid)] += int(cnt)

        return points_by_team

    needed = {"EVENTMSGTYPE", "PERIOD", "SEC_REMAINING", "DESC_UPPER"}
    if not needed.issubset(df_pbp.columns):
        return points_by_team

    df = df_pbp.copy()
    fm = df[
        (df["PERIOD"] == 1) &
        (df["SEC_REMAINING"].notna()) &
        (df["SEC_REMAINING"] > 11 * 60) &
        (df["SEC_REMAINING"] <= 12 * 60)
    ].copy()

    if fm.empty:
        return points_by_team

    def infer_team_id_from_desc(row):
        home = row.get("HOMEDESCRIPTION")
        away = row.get("VISITORDESCRIPTION")
        if isinstance(home, str) and home.strip():
            return home_team_id
        if isinstance(away, str) and away.strip():
            return away_team_id
        return None

    if "PLAYER1_TEAM_ID" in fm.columns:
        fm["TEAM_ID"] = pd.to_numeric(fm["PLAYER1_TEAM_ID"], errors="coerce")
        fm.loc[fm["TEAM_ID"].isna(), "TEAM_ID"] = fm[fm["TEAM_ID"].isna()].apply(infer_team_id_from_desc, axis=1)
    else:
        fm["TEAM_ID"] = fm.apply(infer_team_id_from_desc, axis=1)

    fm = fm[fm["TEAM_ID"].notna()].copy()
    fm["TEAM_ID"] = fm["TEAM_ID"].astype(int)

    fg = fm[fm["EVENTMSGTYPE"] == 1].copy()
    if not fg.empty:
        def fg_points(desc_upper: str) -> int:
            if "3PT" in desc_upper or "3-PT" in desc_upper:
                return 3
            return 2
        fg["PTS"] = fg["DESC_UPPER"].apply(fg_points)
        fg_pts = fg.groupby("TEAM_ID")["PTS"].sum().to_dict()
        for tid, pts in fg_pts.items():
            if tid in points_by_team:
                points_by_team[tid] += int(pts)

    ft = fm[(fm["EVENTMSGTYPE"] == 3) & (~fm["DESC_UPPER"].str.contains("MISS", na=False))].copy()
    if not ft.empty:
        ft_counts = ft.groupby("TEAM_ID").size().to_dict()
        for tid, cnt in ft_counts.items():
            if tid in points_by_team:
                points_by_team[tid] += int(cnt)

    return points_by_team


# ---------------------------------------------------------
#   FIRST SHOT TYPE + FIRST MINUTE METRICS
# ---------------------------------------------------------
def get_team_first_shot_and_first_minute_metrics(game_id: str, home_team_id: int, away_team_id: int) -> pd.DataFrame:
    game_id = str(game_id)
    if is_gleague_game_id(game_id):
        return pd.DataFrame()

    df_pbp = fetch_playbyplay_df(game_id)
    if df_pbp is None or df_pbp.empty:
        return pd.DataFrame()

    home_team_id = int(home_team_id)
    away_team_id = int(away_team_id)

    if {"TEAM_ID", "PERIOD", "SEC_REMAINING"}.issubset(df_pbp.columns):
        df = df_pbp.copy()
        df = df[df["TEAM_ID"].notna()].copy()
        df["TEAM_ID"] = df["TEAM_ID"].astype(int)

        def is_shot_action(row) -> bool:
            if pd.notna(row.get("SHOT_VALUE", None)):
                return True
            at = str(row.get("ACTION_TYPE", "")).lower()
            return "shot" in at or at in ("2pt", "3pt", "jump_shot", "layup", "dunk")

        shots = df[df.apply(is_shot_action, axis=1)].copy()
        if shots.empty:
            return pd.DataFrame()

        if "actionNumber" in shots.columns:
            shots["actionNumber"] = pd.to_numeric(shots["actionNumber"], errors="coerce")
            shots.sort_values(["PERIOD", "actionNumber"], inplace=True)
        else:
            shots.sort_values(["PERIOD"], inplace=True)

        def classify(row) -> str:
            sv = row.get("SHOT_VALUE", None)
            if pd.notna(sv):
                try:
                    return "3PT" if int(sv) == 3 else "2PT"
                except Exception:
                    pass
            d = str(row.get("DESC_UPPER", ""))
            if "3PT" in d or "3-PT" in d:
                return "3PT"
            return "2PT"

        shots["FIRST_SHOT_TYPE"] = shots.apply(classify, axis=1)

        first_shots_by_team = shots.groupby("TEAM_ID").first().reset_index()
        result = first_shots_by_team[["TEAM_ID", "FIRST_SHOT_TYPE"]].copy()
        result["GAME_ID"] = game_id

        # First shot OF GAME -> time for that team only
        result["FIRST_SHOT_TIME_SEC"] = pd.NA
        shots_q1 = shots[(shots["PERIOD"] == 1) & (shots["SEC_REMAINING"].notna())].copy()
        if not shots_q1.empty:
            first_row = shots_q1.iloc[0]
            first_team = int(first_row["TEAM_ID"])
            sec_rem = int(first_row["SEC_REMAINING"])
            sec_rem = max(0, min(12 * 60, sec_rem))
            time_after_tip = (12 * 60) - sec_rem
            result.loc[result["TEAM_ID"] == first_team, "FIRST_SHOT_TIME_SEC"] = int(time_after_tip)

        # FIRST MINUTE BASKETS
        fm = df[
            (df["PERIOD"] == 1) &
            (df["SEC_REMAINING"].notna()) &
            (df["SEC_REMAINING"] >= 11 * 60) &
            (df["SEC_REMAINING"] <= 12 * 60)
        ].copy()

        if fm.empty:
            result["FIRST_MINUTE_BASKETS"] = 0
        else:
            def is_made_shot(row) -> bool:
                sr = str(row.get("SHOT_RESULT", "")).lower()
                if sr in ("made", "make", "1"):
                    return True
                if sr in ("missed", "miss", "0"):
                    return False
                return "MISS" not in str(row.get("DESC_UPPER", ""))

            def is_shot_like(row) -> bool:
                if pd.notna(row.get("SHOT_VALUE", None)):
                    return True
                at = str(row.get("ACTION_TYPE", "")).lower()
                return "shot" in at or at in ("2pt", "3pt", "jump_shot", "layup", "dunk")

            made_fg = fm[fm.apply(is_shot_like, axis=1)].copy()
            if not made_fg.empty:
                made_fg = made_fg[made_fg.apply(is_made_shot, axis=1)].copy()
            fg_counts = made_fg.groupby("TEAM_ID").size().to_dict() if not made_fg.empty else {}

            # Free throw trips (2+ FTs, at least one made) = 1
            ft = fm[fm["DESC_UPPER"].str.contains("FREE THROW", na=False)].copy()
            ft_trip_counts = {}

            if not ft.empty:
                if "actionNumber" in ft.columns:
                    ft["actionNumber"] = pd.to_numeric(ft["actionNumber"], errors="coerce")
                    ft.sort_values(["TEAM_ID", "actionNumber"], inplace=True)
                else:
                    ft.sort_values(["TEAM_ID"], inplace=True)

                for tid, group in ft.groupby("TEAM_ID"):
                    trips = []
                    cur = None
                    for _, row in group.iterrows():
                        desc = str(row.get("DESC_UPPER", ""))
                        m = re.search(r"FREE THROW (\d+) OF (\d+)", desc)
                        if not m:
                            continue
                        x = int(m.group(1))
                        y = int(m.group(2))
                        made = "MISS" not in desc
                        if y < 2:
                            continue

                        if cur is None or x == 1:
                            if cur is not None:
                                trips.append(cur)
                            cur = {"y": y, "made_any": made}
                        else:
                            cur["made_any"] = cur["made_any"] or made

                        if x == y:
                            trips.append(cur)
                            cur = None

                    if cur is not None:
                        trips.append(cur)

                    ft_trip_counts[int(tid)] = sum(1 for t in trips if t["made_any"])

            fmb = {}
            for tid in fm["TEAM_ID"].dropna().unique():
                tid = int(tid)
                fmb[tid] = int(fg_counts.get(tid, 0)) + int(ft_trip_counts.get(tid, 0))

            result["FIRST_MINUTE_BASKETS"] = result["TEAM_ID"].map(fmb).fillna(0).astype(int)

        # FIRST MINUTE POINTS
        pts = compute_first_minute_points(df_pbp, home_team_id, away_team_id)
        result["FIRST_MINUTE_POINTS"] = result["TEAM_ID"].map(pts).fillna(0).astype(int)

        return result[[
            "GAME_ID",
            "TEAM_ID",
            "FIRST_SHOT_TYPE",
            "FIRST_SHOT_TIME_SEC",
            "FIRST_MINUTE_BASKETS",
            "FIRST_MINUTE_POINTS",
        ]]

    return pd.DataFrame()


# ---------------------------------------------------------
#   SCOREBOARD HELPERS
# ---------------------------------------------------------
def get_matchups_for_date(team_id_map, date_str: str) -> pd.DataFrame:
    try:
        sb = with_retries(
            lambda: ScoreboardV2(game_date=date_str),
            label=f"ScoreboardV2({date_str})",
            max_retries=4,
            base_sleep=2.5
        )
        games = sb.get_data_frames()[0]
    except Exception as e:
        print(f"  ❌ ScoreboardV2({date_str}) failed after retries: {e}")
        return pd.DataFrame(columns=["GAME_ID", "Home", "Away"])

    if games is None or games.empty:
        return pd.DataFrame(columns=["GAME_ID", "Home", "Away"])

    required = {"GAME_ID", "HOME_TEAM_ID", "VISITOR_TEAM_ID"}
    if not required.issubset(set(games.columns)):
        return pd.DataFrame(columns=["GAME_ID", "Home", "Away"])

    df = games[["GAME_ID", "HOME_TEAM_ID", "VISITOR_TEAM_ID"]].copy()
    df["GAME_ID"] = df["GAME_ID"].astype(str)
    df = df[~df["GAME_ID"].str.startswith("20")].copy()

    df["Home"] = df["HOME_TEAM_ID"].map(team_id_map)
    df["Away"] = df["VISITOR_TEAM_ID"].map(team_id_map)
    df = df[["GAME_ID", "Home", "Away"]].dropna().copy()

    return dedupe_game_ids(df)


def add_rates_and_combined(df_matchups: pd.DataFrame, agg_filtered: pd.DataFrame) -> pd.DataFrame:
    if df_matchups is None or df_matchups.empty:
        return df_matchups

    out = df_matchups.merge(
        agg_filtered[["Team", "Rate_2PT", "QUIGS_POWER_SCORE"]],
        left_on="Home", right_on="Team", how="left"
    )
    out = out.merge(
        agg_filtered[["Team", "Rate_2PT", "QUIGS_POWER_SCORE"]],
        left_on="Away", right_on="Team", how="left",
        suffixes=("_Home", "_Away")
    )
    out.drop(columns=["Team_Home", "Team_Away"], inplace=True, errors="ignore")

    out["Combined_Rate"] = (out["Rate_2PT_Home"] + out["Rate_2PT_Away"])
    out["Combined_QUIGS_POWER_SCORE"] = (out["QUIGS_POWER_SCORE_Home"] + out["QUIGS_POWER_SCORE_Away"])

    out = round2(out, [
        "Rate_2PT_Home", "Rate_2PT_Away", "Combined_Rate",
        "QUIGS_POWER_SCORE_Home", "QUIGS_POWER_SCORE_Away", "Combined_QUIGS_POWER_SCORE"
    ])

    return out


# ---------------------------------------------------------
#   CACHE LOAD / SAVE
# ---------------------------------------------------------
def load_cache() -> pd.DataFrame:
    base_cols = [
        "GAME_ID",
        "TEAM_ID",
        "FIRST_SHOT_TYPE",
        "FIRST_SHOT_TIME_SEC",
        "FIRST_MINUTE_BASKETS",
        "FIRST_MINUTE_POINTS",
    ]
    if os.path.exists(CACHE_FILE):
        try:
            df = pd.read_csv(CACHE_FILE, dtype={"GAME_ID": str})
            df["GAME_ID"] = df["GAME_ID"].astype(str)
            df = df[~df["GAME_ID"].str.startswith("20")].copy()

            for c in base_cols:
                if c not in df.columns:
                    df[c] = pd.NA

            df["TEAM_ID"] = pd.to_numeric(df["TEAM_ID"], errors="coerce")
            df["FIRST_MINUTE_BASKETS"] = pd.to_numeric(df["FIRST_MINUTE_BASKETS"], errors="coerce").fillna(0).astype(int)
            df["FIRST_MINUTE_POINTS"] = pd.to_numeric(df["FIRST_MINUTE_POINTS"], errors="coerce").fillna(0).astype(int)
            df["FIRST_SHOT_TIME_SEC"] = pd.to_numeric(df["FIRST_SHOT_TIME_SEC"], errors="coerce")

            return df[base_cols].copy()
        except Exception as e:
            print(f"Cache read failed ({CACHE_FILE}): {e}")

    return pd.DataFrame(columns=base_cols)


def save_cache(df: pd.DataFrame) -> None:
    df = df.copy()
    df["GAME_ID"] = df["GAME_ID"].astype(str)
    df = df[~df["GAME_ID"].str.startswith("20")].copy()
    df = df.drop_duplicates(subset=["GAME_ID", "TEAM_ID"], keep="last")
    df.to_csv(CACHE_FILE, index=False)


# ---------------------------------------------------------
#   EXCEL FORMATTING + CHART
# ---------------------------------------------------------
def format_workbook_and_chart(out_path: str):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.chart import LineChart, Reference

    wb = load_workbook(out_path)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # dark
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(vertical="center")
    zebra_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")  # light gray

    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    fmt_2dp = "0.00"

    fmt_pct = "0.00%"
    fmt_date = "yyyy-mm-dd"

    two_dp_headers = {
        "Rate_2PT",
        "First_Minute_Baskets_Per_Game",
        "AVG_Time_FirstShot",
        "QUIGS_POWER_SCORE",
        "Rate_2PT_Home",
        "Rate_2PT_Away",
        "Combined_Rate",
        "QUIGS_POWER_SCORE_Home",
        "QUIGS_POWER_SCORE_Away",
        "Combined_QUIGS_POWER_SCORE",
        "Bin_Low",
        "Bin_High",
        "Bin_Mid",
    }

    three_dp_headers = set()
    pct_headers = {"Win_Prob"}


    int_headers = {
        "Num_2PT",
        "Num_3PT",
        "First_Minute_Baskets",
        "First_Minute_Baskets_Allowed",
        "Games_Played",
        "Games",
        "Wins",
    }

    date_headers = {"Game_Date", "GAME_DATE"}

    cf_qps_sheets = {"TodaysMatchups", "ThisWeeksMatchups", "Last80DaysMatchups"}

    def header_map(ws):
        max_col = ws.max_column
        headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        return {str(h): i + 1 for i, h in enumerate(headers) if h is not None}

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Zebra striping
        for r in range(2, max_row + 1):
            if (r % 2) == 0:
                for c in range(1, max_col + 1):
                    ws.cell(row=r, column=c).fill = zebra_fill

        h2c = header_map(ws)

        # Number formats
        for h, col_idx in h2c.items():
            if h in date_headers:
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = fmt_date
            elif h in two_dp_headers:
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = fmt_2dp
            elif h in three_dp_headers:
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = fmt_3dp
            elif h in pct_headers:
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = fmt_pct
            elif h in int_headers:
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col_idx).number_format = "0"

        # Conditional formatting on Combined_QUIGS_POWER_SCORE
        if ws.title in cf_qps_sheets and "Combined_QUIGS_POWER_SCORE" in h2c and max_row >= 2:
            col_idx = h2c["Combined_QUIGS_POWER_SCORE"]
            col_letter = get_column_letter(col_idx)
            rng = f"{col_letter}2:{col_letter}{max_row}"
            ws.conditional_formatting.add(
                rng,
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B"
                )
            )

        # Win/Loss coloring
        if "Win/Loss" in h2c and max_row >= 2:
            wl_col_letter = get_column_letter(h2c["Win/Loss"])
            wl_range = f"{wl_col_letter}2:{wl_col_letter}{max_row}"
            ws.conditional_formatting.add(wl_range, FormulaRule(formula=[f'{wl_col_letter}2="Win"'], fill=win_fill))
            ws.conditional_formatting.add(wl_range, FormulaRule(formula=[f'{wl_col_letter}2="Loss"'], fill=loss_fill))

        # Auto-fit column widths
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            max_len = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 55)

    # Add chart to BinnedWinRate sheet (if present)
    if "BinnedWinRate" in wb.sheetnames:
        ws = wb["BinnedWinRate"]
        h2c = header_map(ws)
        if ws.max_row >= 2 and "Bin_Mid" in h2c and "Win_Rate" in h2c:
            x_col = h2c["Bin_Mid"]
            y_col = h2c["Win_Rate"]
            max_row = ws.max_row

            # Ensure Win_Rate looks like 0.xx
            for r in range(2, max_row + 1):
                ws.cell(row=r, column=y_col).number_format = "0.00"

            chart = LineChart()
            chart.title = "Binned Win Rate vs Combined_QUIGS_POWER_SCORE"
            chart.y_axis.title = "Win Rate"
            chart.x_axis.title = "Combined QUIGS Power Score (Bin Midpoint)"
            chart.legend = None

            data = Reference(ws, min_col=y_col, min_row=1, max_col=y_col, max_row=max_row)
            cats = Reference(ws, min_col=x_col, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.marker = True

            # Place chart
            ws.add_chart(chart, "I2")

    wb.save(out_path)


# ---------------------------------------------------------
#   MAIN
# ---------------------------------------------------------
def main():
    games_df = get_games_for_season(SEASON, SEASON_TYPE)

    team_id_map = (
        games_df[["TEAM_ID", "TEAM_ABBREVIATION"]]
        .drop_duplicates()
        .set_index("TEAM_ID")["TEAM_ABBREVIATION"]
        .to_dict()
    )

    # Assign home/away teams from MATCHUP
    game_roles = {}
    for gid, group in games_df.groupby("GAME_ID"):
        g = group.copy()
        g["MATCHUP"] = g["MATCHUP"].astype(str)

        home = g[g["MATCHUP"].str.contains("vs.", na=False)]
        away = g[g["MATCHUP"].str.contains("@", na=False)]

        if home.empty or away.empty:
            continue

        game_roles[str(gid)] = {
            "home": int(home.iloc[0]["TEAM_ID"]),
            "away": int(away.iloc[0]["TEAM_ID"]),
        }

    cache_df = load_cache()
    processed_game_ids = set(cache_df["GAME_ID"].astype(str).unique()) if not cache_df.empty else set()

    all_game_ids = set(games_df["GAME_ID"].astype(str).unique())
    new_game_ids = sorted(list(all_game_ids - processed_game_ids))

    game_date_map = (
        games_df[["GAME_ID", "GAME_DATE"]]
        .drop_duplicates()
        .set_index("GAME_ID")["GAME_DATE"]
        .to_dict()
    )

    print(f"\nCache: {len(processed_game_ids)} games already processed.")
    print(f"New games to fetch: {len(new_game_ids)}")

    new_rows = []
    for i, gid in enumerate(new_game_ids, start=1):
        gid_str = str(gid)
        if is_gleague_game_id(gid_str):
            continue

        gdate = game_date_map.get(gid_str, "")
        print(f"[NEW {i}/{len(new_game_ids)}] Processing game {gid_str} ({gdate})...")

        roles = game_roles.get(gid_str)
        if not roles:
            print(f"  Skipping {gid_str}: couldn't determine home/away")
            continue

        try:
            df = get_team_first_shot_and_first_minute_metrics(gid_str, roles["home"], roles["away"])
            if not df.empty:
                new_rows.append(df)
        except Exception as e:
            print(f"  Error processing {gid_str}: {e}")

        time.sleep(SLEEP_BETWEEN_CALLS + random.uniform(0.05, 0.25))

    if new_rows:
        new_df = pd.concat(new_rows, ignore_index=True)
        combined_cache = pd.concat([cache_df, new_df], ignore_index=True)

        combined_cache = combined_cache[[
            "GAME_ID",
            "TEAM_ID",
            "FIRST_SHOT_TYPE",
            "FIRST_SHOT_TIME_SEC",
            "FIRST_MINUTE_BASKETS",
            "FIRST_MINUTE_POINTS",
        ]].copy()

        save_cache(combined_cache)
        cache_df = load_cache()
        print(f"Cache updated: now {cache_df['GAME_ID'].nunique()} games in {CACHE_FILE}")
    else:
        print("No new games added to cache.")

    # Merge abbreviations + date onto cached per-team results
    merged_base = cache_df.copy()
    merged_base["TEAM_ID"] = pd.to_numeric(merged_base["TEAM_ID"], errors="coerce")

    gsmall = games_df[["GAME_ID", "TEAM_ID", "TEAM_ABBREVIATION", "GAME_DATE"]].drop_duplicates()
    gsmall["GAME_ID"] = gsmall["GAME_ID"].astype(str)

    merged = merged_base.merge(gsmall, on=["GAME_ID", "TEAM_ID"], how="left")
    merged.sort_values(["GAME_DATE", "GAME_ID", "TEAM_ABBREVIATION"], inplace=True)

    # -------------------------
    # TEAM BREAKDOWN
    # -------------------------
    agg = (
        merged.groupby("TEAM_ABBREVIATION")["FIRST_SHOT_TYPE"]
        .value_counts()
        .unstack(fill_value=0)
        .rename(columns={"2PT": "Num_2PT", "3PT": "Num_3PT"})
    )
    if "Num_2PT" not in agg.columns:
        agg["Num_2PT"] = 0
    if "Num_3PT" not in agg.columns:
        agg["Num_3PT"] = 0

    fm_tot = merged.groupby("TEAM_ABBREVIATION")["FIRST_MINUTE_BASKETS"].sum()
    gcount = merged.groupby("TEAM_ABBREVIATION")["GAME_ID"].nunique()

    agg = agg.join(fm_tot.rename("First_Minute_Baskets"))
    agg = agg.join(gcount.rename("Games_Played"))

    agg["First_Minute_Baskets"] = agg["First_Minute_Baskets"].fillna(0).astype(int)
    agg["Games_Played"] = agg["Games_Played"].fillna(0).astype(int)

    # First_Minute_Baskets_Allowed (opponent FIRST_MINUTE_BASKETS in same game)
    opp = merged[["GAME_ID", "TEAM_ABBREVIATION", "FIRST_MINUTE_BASKETS"]].copy()
    opp.rename(columns={"TEAM_ABBREVIATION": "OPP_TEAM", "FIRST_MINUTE_BASKETS": "OPP_FIRST_MINUTE_BASKETS"}, inplace=True)

    allowed_df = merged.merge(opp, on="GAME_ID", how="left")
    allowed_df = allowed_df[allowed_df["TEAM_ABBREVIATION"] != allowed_df["OPP_TEAM"]].copy()

    allowed_tot = (
        allowed_df.groupby("TEAM_ABBREVIATION")["OPP_FIRST_MINUTE_BASKETS"]
        .sum()
        .rename("First_Minute_Baskets_Allowed")
    )
    agg = agg.join(allowed_tot)
    agg["First_Minute_Baskets_Allowed"] = agg["First_Minute_Baskets_Allowed"].fillna(0).astype(int)

    # AVG_Time_FirstShot: average only when team had FIRST shot attempt of game
    avg_time_firstshot = merged.groupby("TEAM_ABBREVIATION")["FIRST_SHOT_TIME_SEC"].mean().rename("AVG_Time_FirstShot")
    agg = agg.join(avg_time_firstshot)
    agg["AVG_Time_FirstShot"] = pd.to_numeric(agg["AVG_Time_FirstShot"], errors="coerce")

    denom = (agg["Num_2PT"] + agg["Num_3PT"]).replace({0: pd.NA})
    agg["Rate_2PT_raw"] = agg["Num_2PT"] / denom
    agg["FMBPG_raw"] = agg["First_Minute_Baskets"] / agg["Games_Played"].replace({0: pd.NA})

    allowed_per_game = agg["First_Minute_Baskets_Allowed"] / agg["Games_Played"].replace({0: pd.NA})

    # Your current QPS_raw formula (kept as-is from your code)
    time_cap = (
        pd.to_numeric(agg["AVG_Time_FirstShot"], errors="coerce")
        .fillna(20)
        .clip(lower=0, upper=20)
    )

    agg["QPS_raw"] = -1/(
        (agg["FMBPG_raw"] / 1.6)
        - (time_cap / 12.0)
        + ((agg["FMBPG_raw"] * allowed_per_game) / 3.5)
    )

    agg["Rate_2PT"] = agg["Rate_2PT_raw"].round(2).fillna(0)
    agg["First_Minute_Baskets_Per_Game"] = agg["FMBPG_raw"].round(2).fillna(0)
    agg["AVG_Time_FirstShot"] = agg["AVG_Time_FirstShot"].round(2)
    agg["QUIGS_POWER_SCORE"] = agg["QPS_raw"].round(2).fillna(0)

    agg.reset_index(inplace=True)
    agg.rename(columns={"TEAM_ABBREVIATION": "Team"}, inplace=True)

    agg = agg[
        [
            "Team",
            "Num_2PT",
            "Num_3PT",
            "Rate_2PT",
            "First_Minute_Baskets",
            "First_Minute_Baskets_Allowed",
            "Games_Played",
            "First_Minute_Baskets_Per_Game",
            "AVG_Time_FirstShot",
            "QUIGS_POWER_SCORE",
        ]
    ]

    agg_filtered = agg[agg["Games_Played"] > 10].copy()
    agg_filtered = agg_filtered.sort_values("QUIGS_POWER_SCORE", ascending=False).reset_index(drop=True)

    # -------------------------
    # TODAY'S MATCHUPS
    # -------------------------
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    matchups_base = get_matchups_for_date(team_id_map, today_str)
    matchups = add_rates_and_combined(matchups_base, agg_filtered)
    matchups = dedupe_game_ids(matchups)
    if matchups is not None and not matchups.empty and "Combined_QUIGS_POWER_SCORE" in matchups.columns:
        matchups = matchups.sort_values("Combined_QUIGS_POWER_SCORE", ascending=False).reset_index(drop=True)

    # -------------------------
    # THIS WEEK'S MATCHUPS
    # -------------------------
    week_rows = []
    for i in range(0, 6):
        d_str = (datetime.date.today() + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
        print(f"\nFetching scoreboard for {d_str} (ThisWeeksMatchups) ...")
        day_df = get_matchups_for_date(team_id_map, d_str)
        if day_df is None or day_df.empty:
            continue
        day_df = dedupe_game_ids(day_df)
        day_df.insert(0, "Game_Date", d_str)
        week_rows.append(day_df)
        time.sleep(SCOREBOARD_SLEEP)

    if week_rows:
        this_week_base = pd.concat(week_rows, ignore_index=True)
        this_week_base = this_week_base.drop_duplicates(subset=["GAME_ID"], keep="first").copy()

        tmp = add_rates_and_combined(this_week_base.drop(columns=["Game_Date"]), agg_filtered)
        this_week = tmp.copy()
        this_week.insert(0, "Game_Date", this_week_base["Game_Date"].values)

        this_week = this_week[
            [
                "Game_Date",
                "Home",
                "Away",
                "Rate_2PT_Home",
                "Rate_2PT_Away",
                "Combined_Rate",
                "QUIGS_POWER_SCORE_Home",
                "QUIGS_POWER_SCORE_Away",
                "Combined_QUIGS_POWER_SCORE",
                "GAME_ID",
            ]
        ].copy()

        this_week = this_week.sort_values("Combined_QUIGS_POWER_SCORE", ascending=False).reset_index(drop=True)
    else:
        this_week = pd.DataFrame(columns=[
            "Game_Date", "Home", "Away",
            "Rate_2PT_Home", "Rate_2PT_Away", "Combined_Rate",
            "QUIGS_POWER_SCORE_Home", "QUIGS_POWER_SCORE_Away", "Combined_QUIGS_POWER_SCORE",
            "GAME_ID"
        ])

    # -------------------------
    # LAST 80 DAYS MATCHUPS
    # -------------------------
    today = datetime.date.today()
    cutoff = today - datetime.timedelta(days=80)

    gd = games_df.copy()
    gd["GAME_ID"] = gd["GAME_ID"].astype(str)
    gd = gd[~gd["GAME_ID"].str.startswith("20")].copy()

    gd["GAME_DATE_DT"] = pd.to_datetime(gd["GAME_DATE"], errors="coerce").dt.date
    gd = gd[gd["GAME_DATE_DT"].notna()].copy()
    gd = gd[(gd["GAME_DATE_DT"] >= cutoff) & (gd["GAME_DATE_DT"] < today)].copy()

    game_rows = []
    for gid, group in gd.groupby("GAME_ID"):
        g = group.copy()
        g["MATCHUP"] = g["MATCHUP"].astype(str)
        home = g[g["MATCHUP"].str.contains("vs.", na=False)]
        away = g[g["MATCHUP"].str.contains("@", na=False)]
        if home.empty or away.empty:
            continue

        home_id = int(home.iloc[0]["TEAM_ID"])
        away_id = int(away.iloc[0]["TEAM_ID"])
        game_date = home.iloc[0]["GAME_DATE_DT"]

        game_rows.append({
            "Game_Date": str(game_date),
            "GAME_ID": str(gid),
            "HOME_TEAM_ID": home_id,
            "VISITOR_TEAM_ID": away_id,
            "Home": team_id_map.get(home_id),
            "Away": team_id_map.get(away_id),
        })

    last80 = pd.DataFrame(game_rows)

    if not last80.empty:
        last80 = last80.dropna(subset=["Home", "Away"]).copy()

        last80 = last80.merge(
            agg_filtered[["Team", "Rate_2PT", "QUIGS_POWER_SCORE"]],
            left_on="Home", right_on="Team", how="left"
        )
        last80 = last80.merge(
            agg_filtered[["Team", "Rate_2PT", "QUIGS_POWER_SCORE"]],
            left_on="Away", right_on="Team", how="left",
            suffixes=("_Home", "_Away")
        )
        last80.drop(columns=["Team_Home", "Team_Away"], inplace=True, errors="ignore")

        last80["Combined_Rate"] = (last80["Rate_2PT_Home"] + last80["Rate_2PT_Away"])
        last80["Combined_QUIGS_POWER_SCORE"] = (last80["QUIGS_POWER_SCORE_Home"] + last80["QUIGS_POWER_SCORE_Away"])

        last80 = round2(last80, [
            "Rate_2PT_Home", "Rate_2PT_Away", "Combined_Rate",
            "QUIGS_POWER_SCORE_Home", "QUIGS_POWER_SCORE_Away", "Combined_QUIGS_POWER_SCORE"
        ])

        # Win/Loss from cached FIRST_MINUTE_POINTS
        cache_pts = cache_df[["GAME_ID", "TEAM_ID", "FIRST_MINUTE_POINTS"]].copy()
        cache_pts["GAME_ID"] = cache_pts["GAME_ID"].astype(str)
        cache_pts["TEAM_ID"] = pd.to_numeric(cache_pts["TEAM_ID"], errors="coerce").astype("Int64")
        cache_pts["FIRST_MINUTE_POINTS"] = pd.to_numeric(cache_pts["FIRST_MINUTE_POINTS"], errors="coerce").fillna(0).astype(int)

        def lookup_points(gid_str, team_id_int):
            m = cache_pts[(cache_pts["GAME_ID"] == gid_str) & (cache_pts["TEAM_ID"] == int(team_id_int))]
            if m.empty:
                return None
            return int(m.iloc[0]["FIRST_MINUTE_POINTS"])

        winloss = []
        for _, row in last80.iterrows():
            gid = str(row["GAME_ID"])
            home_id = int(row["HOME_TEAM_ID"])
            away_id = int(row["VISITOR_TEAM_ID"])

            home_pts = lookup_points(gid, home_id)
            away_pts = lookup_points(gid, away_id)

            if home_pts is None or away_pts is None:
                winloss.append("NA")
            else:
                winloss.append("Win" if (home_pts >= 1 and away_pts >= 1) else "Loss")

        last80["Win/Loss"] = winloss

        last80 = last80[
            [
                "Game_Date",
                "Home",
                "Away",
                "Rate_2PT_Home",
                "Rate_2PT_Away",
                "Combined_Rate",
                "QUIGS_POWER_SCORE_Home",
                "QUIGS_POWER_SCORE_Away",
                "Combined_QUIGS_POWER_SCORE",
                "Win/Loss",
                "GAME_ID",
            ]
        ].copy()

        last80["Game_Date"] = pd.to_datetime(last80["Game_Date"], errors="coerce")
        last80.sort_values("Game_Date", ascending=False, inplace=True)
        last80.reset_index(drop=True, inplace=True)
    else:
        last80 = pd.DataFrame(columns=[
            "Game_Date", "Home", "Away",
            "Rate_2PT_Home", "Rate_2PT_Away", "Combined_Rate",
            "QUIGS_POWER_SCORE_Home", "QUIGS_POWER_SCORE_Away", "Combined_QUIGS_POWER_SCORE",
            "Win/Loss", "GAME_ID"
        ])

    # -------------------------
    # BINNED WIN RATE TABLE + WIN PROBABILITIES
    # -------------------------
    binned_table, edges, _labels = compute_binned_win_rates(
        last80,
        score_col="Combined_QUIGS_POWER_SCORE",
        outcome_col="Win/Loss",
        bins=5,
    )

    matchups = add_win_probabilities(matchups, binned_table, edges, score_col="Combined_QUIGS_POWER_SCORE", prob_col="Win_Prob")
    this_week = add_win_probabilities(this_week, binned_table, edges, score_col="Combined_QUIGS_POWER_SCORE", prob_col="Win_Prob")

    # Keep report-like sorting after Win_Prob is added
    if matchups is not None and not matchups.empty and "Combined_QUIGS_POWER_SCORE" in matchups.columns:
        matchups = matchups.sort_values("Combined_QUIGS_POWER_SCORE", ascending=False).reset_index(drop=True)
    if this_week is not None and not this_week.empty and "Combined_QUIGS_POWER_SCORE" in this_week.columns:
        this_week = this_week.sort_values("Combined_QUIGS_POWER_SCORE", ascending=False).reset_index(drop=True)

    # -------------------------
    # SAVE EXCEL (then format + chart)
    # -------------------------
    out = f"first_shot_report_{SEASON.replace('-', '')}.xlsx"

    try:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            agg_filtered.to_excel(writer, sheet_name="TeamBreakdown", index=False)
            matchups.to_excel(writer, sheet_name="TodaysMatchups", index=False)
            this_week.to_excel(writer, sheet_name="ThisWeeksMatchups", index=False)
            last80.to_excel(writer, sheet_name="Last80DaysMatchups", index=False)
            binned_table.to_excel(writer, sheet_name="BinnedWinRate", index=False)

        format_workbook_and_chart(out)
        print("\nSaved Excel:", out)

    except ModuleNotFoundError:
        print("openpyxl missing — writing CSVs instead (formatting + chart not available).")
        agg_filtered.to_csv("TeamBreakdown.csv", index=False)
        matchups.to_csv("TodaysMatchups.csv", index=False)
        this_week.to_csv("ThisWeeksMatchups.csv", index=False)
        last80.to_csv("Last80DaysMatchups.csv", index=False)
        binned_table.to_csv("BinnedWinRate.csv", index=False)

    print("\nToday's matchups:")
    print(matchups)

    print("\nThis week's matchups (today + next 5 days):")
    print(this_week)

    print("\nLast 80 days matchups (excluding today):")
    print(last80.head(30))

    print("\nBinned win rates:")
    print(binned_table)


if __name__ == "__main__":
    main()
